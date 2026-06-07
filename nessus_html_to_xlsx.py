#!/usr/bin/env python3
"""
Nessus HTML Report → Excel (XLSX) converter.

Usage:
    python nessus_html_to_xlsx.py [OPTIONS] report1.html [report2.html ...] output.xlsx

Options:
    --min-risk {critical,high,medium}   Minimum risk level to export (default: medium)
    --split-hosts                       One row per host instead of merged host list
"""

import argparse
import ipaddress
import re
import sys
from collections import defaultdict
from pathlib import Path

import pandas as pd
from bs4 import BeautifulSoup
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── constants ────────────────────────────────────────────────────────────────

ORDER = {"critical": 0, "high": 1, "medium": 2, "low": 3}
OUTPUT_COLUMNS = [
    "Vulnerability Name",
    "Risk Factor",
    "Affected IP:Port",
    "CVSS v3.0 Base Score",
    "CVE",
]

IP_RE    = re.compile(r"\b(?:\d{1,3}\.){3}\d{1,3}\b")
PORT_RE  = re.compile(r"(?:tcp|udp)/(\d{1,5})", re.I)
CVE_RE   = re.compile(r"CVE-\d{4}-\d{4,7}", re.I)
PLUGIN_TITLE_RE = re.compile(r"^\d+\s*-\s*(.+)$")

RISK_COLORS = {
    "critical": "91243E",
    "high":     "DD4B50",
    "medium":   "F18C43",
    "low":      "F8C851",
    "none":     "67ACE1",
}

# ── argument parsing ──────────────────────────────────────────────────────────

def parse_args(argv):
    parser = argparse.ArgumentParser(
        description="Convert Nessus HTML report(s) to a single Excel file."
    )
    parser.add_argument(
        "--min-risk",
        choices=tuple(ORDER),
        default="medium",
        help="Only export findings at or above this risk level (default: medium).",
    )
    parser.add_argument(
        "--split-hosts",
        action="store_true",
        help="One row per host instead of a merged host list.",
    )
    parser.add_argument(
        "paths",
        nargs="+",
        help="Input HTML file(s) followed by the output .xlsx filename.",
    )
    args = parser.parse_args(argv)

    if len(args.paths) < 2:
        parser.error("Provide at least one HTML input and one output .xlsx file.")

    excel_file = args.paths[-1]
    html_files = args.paths[:-1]

    if not excel_file.lower().endswith(".xlsx"):
        parser.error("Output must be a .xlsx file.")

    return html_files, excel_file, args.min_risk, args.split_hosts


# ── helpers ───────────────────────────────────────────────────────────────────

def valid_port(port_text):
    try:
        p = int(port_text)
    except ValueError:
        return False
    return 1 <= p <= 65535


def parse_host_port(host_port):
    host, sep, port = host_port.partition(":")
    port_value = int(port) if sep and port.isdigit() else -1
    try:
        ip_key = int(ipaddress.ip_address(host))
    except ValueError:
        ip_key = float("inf")
    return (ip_key, host, port_value, port)


def sorted_hosts(hosts):
    return sorted(hosts, key=parse_host_port)


def build_accept_set(min_risk):
    threshold = ORDER[min_risk]
    return {risk for risk, rank in ORDER.items() if rank <= threshold}


def warning_path_for(excel_file):
    return str(Path(excel_file).with_suffix(".warnings.txt"))


def add_warning(warnings, stats, message):
    warnings.append(message)
    stats["warning_items"] += 1


def init_stats():
    return {
        "plugins_found": 0,
        "plugins_exported": 0,
        "findings_exported": 0,
        "skipped_no_host": 0,
        "skipped_no_risk": 0,
        "skipped_unknown_risk": 0,
        "skipped_below_min_risk": 0,
        "warning_items": 0,
    }


# ── HTML parsing ──────────────────────────────────────────────────────────────

def _get_detail(wrapper, label):
    """Return the text content of the sibling element following a details-header with the given label."""
    for dh in wrapper.find_all("div", class_="details-header"):
        if dh.get_text(strip=True) == label:
            sib = dh.find_next_sibling()
            return sib.get_text(strip=True) if sib else ""
    return ""


def _extract_ports_from_output_section(wrapper):
    """Extract port numbers from the Plugin Output section of a wrapper."""
    # Find 'Plugin Output' details-header
    for dh in wrapper.find_all("div", class_="details-header"):
        if dh.get_text(strip=True) == "Plugin Output":
            # Scan following h2 tags for tcp/udp port patterns
            ports = set()
            for h2 in wrapper.find_all("h2"):
                for m in PORT_RE.finditer(h2.get_text()):
                    if valid_port(m.group(1)):
                        ports.add(m.group(1))
            return ports or {"0"}
    return {"0"}


def parse_html_file(html_file, vulns, accept, stats, warnings):
    """Parse a single Nessus HTML report and populate vulns dict."""
    try:
        with open(html_file, "r", encoding="utf-8", errors="replace") as f:
            content = f.read()
    except FileNotFoundError:
        print(f"  Error: file not found: {html_file}", file=sys.stderr)
        return False
    except Exception as exc:
        print(f"  Error reading {html_file}: {exc}", file=sys.stderr)
        return False

    soup = BeautifulSoup(content, "lxml")

    # ── Build ordered list of (type, value, element) in document order ──────
    # Host banners: <div style="font-size: 22px; font-weight: 700; ...">IP</div>
    # Plugin headers: <div onclick="toggleSection(...)">
    all_elements = []
    for div in soup.find_all("div"):
        style  = div.get("style", "")
        onclick = div.get("onclick", "")
        div_id  = div.get("id", "")

        if "font-size: 22px" in style and "font-weight: 700" in style:
            text = div.get_text(strip=True)
            ips = IP_RE.findall(text)
            if ips:
                all_elements.append(("host", ips[0], div))

        elif "toggleSection" in onclick and div_id and not div_id.endswith("-container"):
            all_elements.append(("plugin_header", div_id, div))

    # ── Walk elements, tracking current host ─────────────────────────────────
    current_host = None
    total_plugins = sum(1 for t, _, _ in all_elements if t == "plugin_header")
    plugin_index  = 0
    stats["plugins_found"] += total_plugins

    for element_type, value, element in all_elements:

        if element_type == "host":
            current_host = value
            continue

        # element_type == "plugin_header"
        plugin_index += 1
        if plugin_index % 20 == 0 or plugin_index == total_plugins:
            print(f"  Processing plugin {plugin_index}/{total_plugins}", end="\r")

        # ── get the -container sibling ────────────────────────────────────
        container_id = value + "-container"
        wrapper = soup.find("div", id=container_id)
        if wrapper is None:
            stats["skipped_no_risk"] += 1
            add_warning(warnings, stats, f"No container found for plugin id={value}")
            continue

        # ── vulnerability name from the header div ─────────────────────────
        raw_title = element.get_text(strip=True)
        m = PLUGIN_TITLE_RE.match(raw_title)
        vuln_name = m.group(1).strip() if m else raw_title.strip()
        # Strip trailing '-' toggle button text
        vuln_name = vuln_name.rstrip("- ").strip()

        # ── host check ─────────────────────────────────────────────────────
        if not current_host:
            stats["skipped_no_host"] += 1
            add_warning(warnings, stats, f"Missing host before plugin '{vuln_name}'")
            continue

        # ── risk ──────────────────────────────────────────────────────────
        risk_text = _get_detail(wrapper, "Risk Factor").lower()
        if not risk_text:
            stats["skipped_no_risk"] += 1
            add_warning(warnings, stats, f"Missing risk for plugin '{vuln_name}' on {current_host}")
            continue

        risk = risk_text.strip()
        if risk not in ORDER:
            if risk in ("none", "info", "informational"):
                stats["skipped_below_min_risk"] += 1
            else:
                stats["skipped_unknown_risk"] += 1
                add_warning(warnings, stats, f"Unrecognized risk '{risk}' for '{vuln_name}' on {current_host}")
            continue

        if risk not in accept:
            stats["skipped_below_min_risk"] += 1
            continue

        # ── CVSS v3 ───────────────────────────────────────────────────────
        cvss_raw = _get_detail(wrapper, "CVSS v3.0 Base Score")
        cvss_m   = re.match(r"[\d.]+", cvss_raw)
        cvss     = cvss_m.group() if cvss_m else ""

        # ── CVEs ──────────────────────────────────────────────────────────
        refs_text = _get_detail(wrapper, "References")
        cves = [c.upper() for c in CVE_RE.findall(refs_text)]

        # ── ports ─────────────────────────────────────────────────────────
        ports = _extract_ports_from_output_section(wrapper)

        # ── accumulate into vulns ─────────────────────────────────────────
        current = vulns[vuln_name]
        for cve in cves:
            current["cve"].add(cve)

        if not current["risk"] or ORDER[risk] < ORDER[current["risk"]]:
            current["risk"] = risk

        if cvss:
            try:
                if not current["cvss"] or float(cvss) > float(current["cvss"]):
                    current["cvss"] = cvss
            except ValueError:
                pass

        for port in ports:
            current["hosts"].add(f"{current_host}:{port}")

        stats["plugins_exported"] += 1

    print()
    return True


# ── row building ──────────────────────────────────────────────────────────────

def build_rows(vulns, split_hosts=False):
    rows = []
    for name, data in vulns.items():
        risk      = data["risk"].lower()
        host_list = sorted_hosts(data["hosts"])

        if split_hosts:
            for host in host_list:
                rows.append({
                    "Vulnerability Name":  name,
                    "Risk Factor":         risk.title(),
                    "Affected IP:Port":    host,
                    "CVSS v3.0 Base Score": data["cvss"],
                    "CVE":                 ", ".join(sorted(data["cve"])),
                })
        else:
            rows.append({
                "Vulnerability Name":  name,
                "Risk Factor":         risk.title(),
                "Affected IP:Port":    ", ".join(host_list),
                "CVSS v3.0 Base Score": data["cvss"],
                "CVE":                 ", ".join(sorted(data["cve"])),
            })

    rows.sort(key=lambda r: (
        ORDER.get(r["Risk Factor"].lower(), len(ORDER)),
        r["Vulnerability Name"].lower(),
        parse_host_port(r["Affected IP:Port"].split(", ", 1)[0] or "0.0.0.0:0"),
    ))
    return rows


# ── Excel output ──────────────────────────────────────────────────────────────

RISK_HEX = {
    "Critical": "91243E",
    "High":     "DD4B50",
    "Medium":   "F18C43",
    "Low":      "F8C851",
}
HEADER_BG  = "2B4590"
HEADER_FG  = "FFFFFF"


def write_excel(df, excel_file):
    with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Findings")
        ws = writer.sheets["Findings"]

        # Freeze header row
        ws.freeze_panes = "A2"

        # Style header row
        header_fill = PatternFill("solid", start_color=HEADER_BG)
        for cell in ws[1]:
            cell.font      = Font(bold=True, color=HEADER_FG, name="Arial", size=11)
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Style data rows
        risk_col_idx = OUTPUT_COLUMNS.index("Risk Factor") + 1  # 1-based
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            risk_cell = row[risk_col_idx - 1]
            risk_val  = str(risk_cell.value) if risk_cell.value else ""
            bg_hex    = RISK_HEX.get(risk_val, "FFFFFF")
            risk_fill = PatternFill("solid", start_color=bg_hex)
            for cell in row:
                cell.font      = Font(name="Arial", size=10)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            risk_cell.fill = risk_fill
            risk_cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)

        # Auto-fit column widths (capped at 80)
        for col_idx, col_cells in enumerate(ws.iter_cols(), start=1):
            max_len = max(
                (len(str(c.value)) if c.value is not None else 0) for c in col_cells
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 80)

        # Set row height for header
        ws.row_dimensions[1].height = 30


# ── warnings ──────────────────────────────────────────────────────────────────

def write_warnings(warnings, excel_file):
    if not warnings:
        return None
    warning_file = warning_path_for(excel_file)
    with open(warning_file, "w", encoding="utf-8") as f:
        f.write("\n".join(warnings) + "\n")
    return warning_file


# ── main ──────────────────────────────────────────────────────────────────────

def main(argv=None):
    html_files, excel_file, min_risk, split_hosts = parse_args(argv or sys.argv[1:])

    vulns    = defaultdict(lambda: {"risk": "", "cvss": "", "hosts": set(), "cve": set()})
    stats    = init_stats()
    warnings = []
    accept   = build_accept_set(min_risk)

    print(f"\nTotal HTML files : {len(html_files)}")
    print(f"Min risk         : {min_risk.upper()}")
    print(f"Split hosts      : {'ON' if split_hosts else 'OFF'}")

    for idx, html_file in enumerate(html_files, start=1):
        print(f"\n[{idx}/{len(html_files)}] Reading: {html_file}")
        ok = parse_html_file(html_file, vulns, accept, stats, warnings)
        if not ok:
            return 1

    print()
    rows = build_rows(vulns, split_hosts=split_hosts)
    stats["findings_exported"] = len(rows)
    df = pd.DataFrame(rows, columns=OUTPUT_COLUMNS)

    print("Creating Excel file …")
    try:
        write_excel(df, excel_file)
    except Exception as exc:
        print(f"Error writing Excel file: {exc}", file=sys.stderr)
        return 1

    warning_file = write_warnings(warnings, excel_file)

    print("Done.")
    print(f"File             : {excel_file}")
    print(f"Merged from      : {len(html_files)} HTML file(s)")
    print(f"Order            : CRITICAL → HIGH → MEDIUM")
    print(f"Plugins found    : {stats['plugins_found']}")
    print(f"Plugins exported : {stats['plugins_exported']}")
    print(f"Rows exported    : {stats['findings_exported']}")
    skipped = (stats["skipped_no_host"] + stats["skipped_no_risk"] +
               stats["skipped_unknown_risk"] + stats["skipped_below_min_risk"])
    print(f"Plugins skipped  : {skipped}")
    print(f"  Missing host   : {stats['skipped_no_host']}")
    print(f"  Missing risk   : {stats['skipped_no_risk']}")
    print(f"  Unknown risk   : {stats['skipped_unknown_risk']}")
    print(f"  Below min risk : {stats['skipped_below_min_risk']}")
    print(f"Warnings         : {stats['warning_items']}")
    if warning_file:
        print(f"Warnings file    : {warning_file}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
